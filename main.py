import np as np
import openpyxl
import tkinter.filedialog
import statistics
import scipy

#from matplotlib import pyplot as plt

#from matplotlib.widgets import Button
#import mpl_interactions.ipyplot as iplt





import pandas as pd  # (version 1.0.0)
import plotly  # (version 4.5.0)
import plotly.express as px

import dash  # (version 1.8.0)
from dash import dcc
from dash import html
from dash.dependencies import Input, Output, State






#import mpl_interactions.ipyplot as iplt
import mpl_interactions.ipyplot as iplt
import matplotlib.pyplot as plt
import numpy as np

#x = np.linspace(0, np.pi, 100)
#tau = np.linspace(0.5, 10, 100)

#def f1(x, tau, beta):
#    return np.sin(x * tau) * x * beta
#def f2(x, tau, beta):
#    return np.sin(x * beta) * x * tau


#fig, ax = plt.subplots()
#controls = iplt.plot(x, f1, tau=tau, beta=(1, 10, 100), label="f1")
#iplt.plot(x, f2, controls=controls, label="f2")
#_ = plt.legend()
#plt.show()






"""

x = np.linspace(0, np.pi, 100)
p_val = np.linspace(0,10,100)

def graph1(x, p_val, mu):
    return x * p_val * mu

#test

# ACCEPTANCE CRITERA
# AC01: if p value, have a list of proteins fitting the p value
# AC02: bar graph of each individual sample to see variation
# AC03: add typable p value


# NEW FUNCTIONALITY
def proteinMeanGraph(protein_name, mu):
    fig, ax = plt.subplots()
    controls = iplt.plot(x, graph1, p_val=p_val, mu = mu, label=protein_name)
    #iplt.plot(x, graph1, controls=controls)

    _ = plt.legend()

    plt.show()


####################







# @params:
# choice: whether user wants to display graphical representation of results or not
# protein_name: name of protein
# desired_p_val: the value that the user entered for the desired p-value (i.e. 0.05, 0.01)
# mu: population mean


# TODO: improve p-value function
def pvalue_101(choice, protein_name, desired_p_val, mu, sigma, samp_size, samp_mean=0, deltam=0):
    np.random.seed(1234)
    s1 = np.random.normal(mu, sigma, samp_size)  # simulate random normal distribution
    if samp_mean > 0:
        print(len(s1[s1>samp_mean]))
        outliers = float(len(s1[s1>samp_mean])*100)/float(len(s1))
        print('Percentage of numbers larger than {} is {}%'.format(samp_mean, outliers))
    if deltam == 0:
        deltam = abs(mu-samp_mean)
    if deltam > 0 :
        outliers = (float(len(s1[s1>(mu+deltam)]))
                    +float(len(s1[s1<(mu-deltam)])))*100.0/float(len(s1))
        #print('Percentage of numbers further than the population mean of {} by +/-{} is {}% for {}'.format(mu, deltam, outliers, protein_name))
        print('Percentage of numbers significantly larger than the control mean is {}% for {}'.format(outliers, protein_name))

    percentage_needed = 1 - desired_p_val
    if outliers < percentage_needed and choice == "y" and samp_size > 5:
        fig, ax = plt.subplots(figsize=(8,8))
        #fig.suptitle('Normal Distribution: population_mean={}'.format(mu) )
        fig.suptitle('Normal Distribution: population_mean for {}'.format(protein_name))
        plt.hist(s1)
        plt.axvline(x=sigma, color='red')
        #plt.axvline(x=mu-deltam, color='green')
        plt.show()

    return outliers
"""


# Set-up program
# File selection



root = tkinter.Tk()
root.withdraw
fileName1 = tkinter.filedialog.askopenfilename()

dataFrame = openpyxl.load_workbook(fileName1)

dataFrameReader = dataFrame.active

# Creating the data structures
arr = [] # goes down each column
sampleData = {} # goes across each row


sampleName = ""
flag = True
counter = 0
controlSampleNumbers = [] #column number of control samples
covidSampleNumbers = [] #column number of covid samples
covid2wkSampleNumbers = [] #column number of covid 2 week samples
covid6wkSampleNumbers = [] #column number of covid 6 week samples

for col in range(5, dataFrameReader.max_column):
    arr.append([])
    for row in dataFrameReader.iter_rows(0, dataFrameReader.max_row):
        if flag:
            sampleName = row[col].value
            if sampleName.find("ctrl") != -1:
                controlSampleNumbers.append(col)
            if sampleName.find("Cov") != -1:
                covidSampleNumbers.append(col)
                if sampleName.find("2wk") != -1:
                    covid2wkSampleNumbers.append(col)
                if sampleName.find("6wk") != -1:
                    covid6wkSampleNumbers.append(col)
            flag = False

        value = row[col].value
        if not isinstance(value, str):
            arr[counter].append(value)

    arr[counter].sort()
    print("Number of proteins with valid data in", sampleName, "=", len(arr[counter]))
    counter += 1
    flag = True

rowCounter = 0


# Set-up control values
# Goal: parse through just the control column and store data in dictionary
controlDataMap = {}
for row in dataFrameReader.iter_rows(1, dataFrameReader.max_row):
    proteinName = row[2].value
    controlDataMap[proteinName] = []
    for col in controlSampleNumbers:
        value = row[col].value
        if not isinstance(value, str):  # Filtered as value
            controlDataMap[proteinName].append(value)
        else:
            controlDataMap[proteinName].append(50)

controlDataMappedToMean = {}
for key,val in controlDataMap.items():
    if len(val) <= 0:
        continue
    mean = statistics.mean(val)
    controlDataMappedToMean[key] = mean

# get data from left --> right
# key = proteinName
# value = list of data points


# Collect Sample Data
proteinName_mapped_to_p_value = {}
proteinName_mapped_to_mean = {}
numSamples_noCoV = len(controlSampleNumbers)
numSamples_coV = len(covidSampleNumbers)
numSamples_coV2wk = len(covid2wkSampleNumbers)
numSamples_coV6wk = len(covid6wkSampleNumbers)

for row in dataFrameReader.iter_rows(1, dataFrameReader.max_row):
    proteinName = row[2].value
    sampleData[proteinName] = []

    coV_mean = 0
    noCoV_mean = 0
    coV2wk_mean = 0
    coV6wk_mean = 0

    coV_sum = 0
    noCoV_sum = 0
    coV2wk_sum = 0
    coV6wk_sum = 0

    ctrl_count = []
    coV_count = []
    coV2wk_count = []
    coV6wk_count = []
    count = 1

    #Switch commented out parts to evaluate ctrl vs cov, cov2wks vs cov6wks
    for col in range(5, dataFrameReader.max_column): # 6 = start of non-control groups
        value = row[col].value
        if not isinstance(value, str):
            sampleData[proteinName].append(value)
            """
            if col in controlSampleNumbers:
                noCoV_sum += value
                ctrl_count.append(count)
                count += 1
            elif col in covidSampleNumbers:
                coV_sum += value
                coV_count.append(count)
                count += 1
            """
            #"""
            if col in covid2wkSampleNumbers:
                coV2wk_sum += value
                coV2wk_count.append(count)
                count += 1
            if col in covid6wkSampleNumbers:
                coV6wk_sum += value
                coV6wk_count.append(count)
                count += 1
            #"""
        else:
            sampleData[proteinName].append(0)
            """
            if col in controlSampleNumbers:
                noCoV_sum += 0
                ctrl_count.append(count)
                count += 1
            elif col in covidSampleNumbers:
                coV_sum += 0
                coV_count.append(count)
                count += 1
            """
            #"""
            if col in covid2wkSampleNumbers:
                coV2wk_sum += 0
                coV2wk_count.append(count)
                count += 1
            if col in covid6wkSampleNumbers:
                coV6wk_sum += 0
                coV6wk_count.append(count)
                count += 1
            #"""

    noCoV_mean = noCoV_sum / numSamples_noCoV
    coV_mean = coV_sum / numSamples_coV
    coV2wk_mean = coV2wk_sum / numSamples_coV2wk
    coV6wk_mean = coV6wk_sum / numSamples_coV6wk

    #t-test and p-value mapping
    a = []
    b = []
    """
    for col in ctrl_count:
        a.append(sampleData[proteinName][col])
    for col in coV_count:
        b.append(sampleData[proteinName][col])
    """
    #'''
    for col in coV2wk_count:
        a.append(sampleData[proteinName][col])
    for col in coV6wk_count:
        b.append(sampleData[proteinName][col])
    #'''
    res = scipy.stats.ttest_ind(a, b, equal_var=True)
    proteinName_mapped_to_p_value[proteinName] = res.pvalue
    proteinName_mapped_to_mean[proteinName] = "Non-Covid mean: " + str(noCoV_mean) + " | Covid mean: " + str(coV_mean) + " || p-value: " + str(res.pvalue)
    #proteinName_mapped_to_mean[proteinName] = "Covid 2 wk mean: " + str(CoV2wk_mean) + " | Covid 6 wk mean: " + str(coV6wk_mean) + " || p-value: " + str(res.pvalue)

    if numSamples_coV == 0 and numSamples_noCoV == 0:
        proteinName_mapped_to_p_value[proteinName] = 0.99
        proteinName_mapped_to_mean[proteinName] = "Insufficient Data for statistical significance"

    rowCounter += 1



# Get Sample Names
sampleDataListSampleNames = []
ctrlSamples = []
coVSamples = []
coV2wkSamples = []
coV6wkSamples = []
colorMap = {}
for row in dataFrameReader.iter_rows(0, 1):
    for col in range(5, dataFrameReader.max_column):
        s = str(row[col].value)
        #s = s.replace(".PG.Quantity", "")
        s = s.split('_')[0]
        sampleDataListSampleNames.append(s)

        if "ctrl" in s:
            colorMap[s] = "blue"
            ctrlSamples.append(s)
        if "Cov" in s:
            colorMap[s] = "red"
            coVSamples.append(s)
        if "2wk" in s:
            colorMap[s] = "purple"
            coV2wkSamples.append(s)
        if "6wk" in s:
            colorMap[s] = "pink"
            coV6wkSamples.append(s)

#sampleDataListCompare = [ctrlSamples, coVSamples]
sampleDataListCompare = [coV2wkSamples, coV6wkSamples]


proteinSampleDataMappedToMean = {}
#desired_p_value = (float)(input("Enter the desired p value: "))
desired_p_value = 0.05
#graphs = input("Would you like to see a graphical display of the statistics for statistically significant proteins? (y/n)")
graphs = "y"

temp_counter = 0
for key,val in sampleData.items():
    # If entire data set is invalid (i.e. "Filtered) then ignore this sample
    #if temp_counter >= 1:
       # break
    #temp_counter += 1

    if len(val) <= 0:
        continue

    proteinSampleDataMappedToMean[key] = []
    proteinSampleDataMappedToMean[key].append(statistics.mean(val))


    #proteinSampleDataMappedToMean[key].append(pvalue_101(graphs, key, desired_p_value, controlDataMappedToMean[key], (int)(proteinSampleDataMappedToMean[key][0]), 100))


    # proteinSampleDataMappedToMean[key].append(proteinMeanGraph(key, controlDataMappedToMean[key])) # Note: this allows GRAPHING


























print("-----------------------------------------------------------")



proteinNamesList = list(sampleData.keys())

from dash import Dash, dcc, html, Input, Output
import plotly.express as px

app = Dash(__name__)


# List of elements on HTML webpage
app.layout = html.Div([
    html.H2('Covid vs. Non-Covid Protein Expression Application || Property of the Gong Laboratory.'),
    dcc.Dropdown(
        id="dropdown",
        options=[x for x in sampleData.keys()],
        value=proteinNamesList[0],
        clearable=False,
    ),
    html.Label("Include results only from proteins with a p-value of less than or equal to: "),
    dcc.Slider(0, 1, marks=None, value=0.05, id='slider', tooltip={"placement": "bottom", "always_visible": True}),
    html.Label("Manually enter a p-value: "),
    dcc.Input(type='number', value=0.05, id='manualinput'),
    html.H3('Bar Chart'),
    dcc.Graph(id="graph"),
    html.H3('Violin Plot'),
    dcc.Graph(id="violin"),
    html.H3('Density Heatmap'),
    dcc.Graph(id="heatmap"),
    html.H3('Empirical Cumulative Distribution Plot'),
    dcc.Graph(id="ECDF")
])

@app.callback(
    Output("dropdown", "options"),
    Input("slider", "value"))
def update_dropdown(pVal):
    return [key for key,val in proteinName_mapped_to_p_value.items() if val <= pVal]


@app.callback(
    Output("slider", "value"),
    Input("manualinput", "value"))
def update_proteins(pVal):
    if pVal is None:
        return 0
    return pVal

@app.callback(
    Output("graph", "figure"),
    Input("dropdown", "value"))
def update_bar_chart(proteinName):
    print(proteinName + ": " + proteinName_mapped_to_mean[proteinName])
    df = {'Samples': sampleDataListSampleNames, 'Detection Levels': sampleData[proteinName]}
    fig = px.bar(df, x = 'Samples', y='Detection Levels', color="Samples", color_discrete_map=colorMap)
    return fig

@app.callback(
    Output("violin", "figure"),
    Input("dropdown", "value"))
def update_violin_plot(proteinName):
    df = {'Detection Levels': sampleData[proteinName], 'Sample': sampleDataListSampleNames}
    fig = px.violin(df, y='Detection Levels', hover_data=['Sample'], color_discrete_map=colorMap, box=True, points="all")
    return fig

@app.callback(
    Output("heatmap", "figure"),
    Input("dropdown", "value"))
def update_heatmap_plot(proteinName):
    df = {'Samples': sampleDataListSampleNames, 'Detection Levels': sampleData[proteinName]}
    fig = px.density_heatmap(df, y='Detection Levels', x='Samples', marginal_x="histogram", marginal_y="histogram")
    return fig

@app.callback(
    Output("ECDF", "figure"),
    Input("dropdown", "value"))
def update_ECDF_plot(proteinName):
    df = {'Detection Levels': sampleData[proteinName], 'Sample': sampleDataListSampleNames}
    fig = px.ecdf(df, x='Detection Levels', hover_data=['Sample'], markers = True)
    return fig

app.run_server(debug=True)
























"""
app = dash.Dash(__name__)

# ---------------------------------------------------------------

df = pd.read_csv(
    "C:\\Users\\antho\\Downloads\\DOHMH_New_York_City_Restaurant_Inspection_Results.csv")  # https://drive.google.com/file/d/1jyvSiRjaNIeOCP59dUFQuZ0_N_StiQOr/view?usp=sharing
df['INSPECTION DATE'] = pd.to_datetime(df['INSPECTION DATE'])
df = df.groupby(['INSPECTION DATE', 'CUISINE DESCRIPTION', 'CAMIS'], as_index=False)['SCORE'].mean()
df = df.set_index('INSPECTION DATE')
df = df.loc['2016-01-01':'2019-12-31']
df = df.groupby([pd.Grouper(freq="M"), 'CUISINE DESCRIPTION'])['SCORE'].mean().reset_index()
# print (df[:5])




# ---------------------------------------------------------------
app.layout = html.Div([

    html.Div([
        dcc.Graph(id='our_graph')
    ], className='nine columns'),

    html.Div([

        html.Br(),
        html.Label(['Choose 3 Cuisines to Compare:'], style={'font-weight': 'bold', "text-align": "center"}),
        dcc.Dropdown(id='cuisine_one',
                     options=[{'label': x, 'value': x} for x in
                              proteinSampleDataMappedToMean.keys()],
                     value='African',
                     multi=False,
                     disabled=False,
                     clearable=True,
                     searchable=True,
                     placeholder='Choose Cuisine...',
                     className='form-dropdown',
                     style={'width': "90%"},
                     persistence='string',
                     persistence_type='memory'),

        dcc.Dropdown(id='cuisine_two',
                     options=[{'label': x, 'value': x} for x in
                              df.sort_values('CUISINE DESCRIPTION')['CUISINE DESCRIPTION'].unique()],
                     value='Asian',
                     multi=False,
                     clearable=False,
                     persistence='string',
                     persistence_type='session'),

        dcc.Dropdown(id='cuisine_three',
                     options=[{'label': x, 'value': x} for x in
                              df.sort_values('CUISINE DESCRIPTION')['CUISINE DESCRIPTION'].unique()],
                     value='Donuts',
                     multi=False,
                     clearable=False,
                     persistence='string',
                     persistence_type='local'),

    ], className='three columns'),

])


# ---------------------------------------------------------------

@app.callback(
    Output('our_graph', 'figure'),
    [Input('cuisine_one', 'value'),
     Input('cuisine_two', 'value'),
     Input('cuisine_three', 'value')]
)
def build_graph(first_cuisine, second_cuisine, third_cuisine):


    dff = df[(df['CUISINE DESCRIPTION'] == first_cuisine) |
             (df['CUISINE DESCRIPTION'] == second_cuisine) |
             (df['CUISINE DESCRIPTION'] == third_cuisine)]
    # print(dff[:5])

    fig = px.line(dff, x="INSPECTION DATE", y="SCORE", color='CUISINE DESCRIPTION', height=600)
    fig.update_layout(yaxis={'title': 'NEGATIVE POINT'},
                      title={'text': 'Restaurant Inspections in NYC',
                             'font': {'size': 28}, 'x': 0.5, 'xanchor': 'center'})
    return fig


# ---------------------------------------------------------------

if __name__ == '__main__':
    app.run_server(debug=False)

"""










"""
app = dash.Dash(__name__)

# ---------------------------------------------------------------

df = pd.read_csv(
    "C:\\Users\\antho\\Downloads\\DOHMH_New_York_City_Restaurant_Inspection_Results.csv")  # https://drive.google.com/file/d/1jyvSiRjaNIeOCP59dUFQuZ0_N_StiQOr/view?usp=sharing
df['INSPECTION DATE'] = pd.to_datetime(df['INSPECTION DATE'])
df = df.groupby(['INSPECTION DATE', 'CUISINE DESCRIPTION', 'CAMIS'], as_index=False)['SCORE'].mean()
df = df.set_index('INSPECTION DATE')
df = df.loc['2016-01-01':'2019-12-31']
df = df.groupby([pd.Grouper(freq="M"), 'CUISINE DESCRIPTION'])['SCORE'].mean().reset_index()
# print (df[:5])

df = px.data.iris()





# ---------------------------------------------------------------
app.layout = html.Div([

    html.Div([
        dcc.Graph(id='our_graph')
    ], className='nine columns'),

    html.Div([

        html.Br(),
        html.Label(['Choose 3 Cuisines to Compare:'], style={'font-weight': 'bold', "text-align": "center"}),
        dcc.Dropdown(id='cuisine_one',
                     options=[{'label': x, 'value': x} for x in
                              proteinSampleDataMappedToMean.keys()],
                     value='African',
                     multi=False,
                     disabled=False,
                     clearable=True,
                     searchable=True,
                     placeholder='Choose Cuisine...',
                     className='form-dropdown',
                     style={'width': "90%"},
                     persistence='string',
                     persistence_type='memory'),

        dcc.Dropdown(id='cuisine_two',
                     options=[{'label': x, 'value': x} for x in
                              df.sort_values('CUISINE DESCRIPTION')['CUISINE DESCRIPTION'].unique()],
                     value='Asian',
                     multi=False,
                     clearable=False,
                     persistence='string',
                     persistence_type='session'),

        dcc.Dropdown(id='cuisine_three',
                     options=[{'label': x, 'value': x} for x in
                              df.sort_values('CUISINE DESCRIPTION')['CUISINE DESCRIPTION'].unique()],
                     value='Donuts',
                     multi=False,
                     clearable=False,
                     persistence='string',
                     persistence_type='local'),

    ], className='three columns'),

])


# ---------------------------------------------------------------

@app.callback(
    Output('our_graph', 'figure'),
    [Input('cuisine_one', 'value'),
     Input('cuisine_two', 'value'),
     Input('cuisine_three', 'value')]
)
def build_graph(first_cuisine, second_cuisine, third_cuisine):


   # dff = df[(df['CUISINE DESCRIPTION'] == first_cuisine) |
             #(df['CUISINE DESCRIPTION'] == second_cuisine) |
             #(df['CUISINE DESCRIPTION'] == third_cuisine)]

    fig = px.scatter(df, x='sepal_length', y='sepal_width', color='species', size='petal_length')
    return fig


# ---------------------------------------------------------------

if __name__ == '__main__':
    app.run_server(debug=False)
"""




"""
from dash import Dash, dcc, html
from base64 import b64encode
import io

app = Dash(__name__)

buffer = io.StringIO()

df = px.data.iris() # replace with your own data source
fig = px.scatter(
    df, x="sepal_width", y="sepal_length",
    color="species")
fig.write_html(buffer)

html_bytes = buffer.getvalue().encode()
encoded = b64encode(html_bytes).decode()

app.layout = html.Div([
    html.H4('Simple plot export options'),
    html.P("↓↓↓ try downloading the plot as PNG ↓↓↓", style={"text-align": "right", "font-weight": "bold"}),
    dcc.Graph(id="graph", figure=fig),
    html.A(
        html.Button("Download as HTML"),
        id="download",
        href="data:text/html;base64," + encoded,
        download="plotly_graph.html"
    )
])


app.run_server(debug=True)
"""


